#!/usr/bin/env python3
"""Query simple Excel workbooks for tabular data."""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_LIMIT = 10
SALARY_ALIASES = [
    "\u6708\u85aa",
    "\u85aa\u8d44",
    "\u5de5\u8d44",
    "\u85aa\u916c",
    "salary",
    "monthly salary",
    "monthly_salary",
]


def resolve_file_path(raw_path: str) -> Path:
    """Resolve user or WebUI paths against the project root."""
    if not isinstance(raw_path, str) or not raw_path.strip():
        raise ValueError("--file must be a non-empty path")

    normalized = raw_path.strip().strip('"').strip("'")
    normalized_slashes = normalized.replace("\\", "/")

    if normalized_slashes.startswith("/uploads/"):
        candidate = PROJECT_ROOT / normalized_slashes.lstrip("/")
    else:
        candidate = Path(normalized)
        if not candidate.is_absolute():
            candidate = PROJECT_ROOT / candidate

    candidate = candidate.resolve()
    if candidate.exists() and candidate.is_file():
        return candidate

    fallback = PROJECT_ROOT / "uploads" / Path(normalized_slashes).name
    if fallback.exists() and fallback.is_file():
        return fallback.resolve()

    raise FileNotFoundError(f"Excel file not found: {candidate}")


def load_sheet(path: Path, sheet_name: str | None) -> tuple[str, list[list[Any]]]:
    """Load a worksheet by name or use the active sheet."""
    with path.open("rb") as handle:
        workbook = load_workbook(handle, read_only=True, data_only=True)

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                available = ", ".join(workbook.sheetnames)
                raise ValueError(f"Sheet not found: {sheet_name}. Available: {available}")
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active

        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        return worksheet.title, rows


def build_records(rows: list[list[Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    """Convert worksheet rows into dict records using the first non-empty row."""
    header_index = None
    headers: list[str] = []

    for index, row in enumerate(rows):
        if not any(not is_blank(value) for value in row):
            continue
        header_index = index
        headers = normalize_headers(row)
        break

    if header_index is None:
        return [], []

    records: list[dict[str, Any]] = []
    for excel_row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not any(not is_blank(value) for value in row):
            continue

        record = {
            headers[index]: row[index] if index < len(row) else None
            for index in range(len(headers))
        }
        record["_excel_row"] = excel_row_number
        records.append(record)

    return headers, records


def normalize_headers(row: Iterable[Any]) -> list[str]:
    """Create stable non-empty headers and de-duplicate repeated names."""
    headers: list[str] = []
    counts: dict[str, int] = {}

    for index, value in enumerate(row, start=1):
        raw_header = "" if value is None else str(value).strip()
        header = raw_header or f"Column{index}"
        count = counts.get(header, 0) + 1
        counts[header] = count
        if count > 1:
            header = f"{header}_{count}"
        headers.append(header)

    return headers


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def normalize_key(value: str) -> str:
    return re.sub(r"[\s_\-()/（）]+", "", str(value).strip().lower())


def find_column(headers: list[str], requested: str | None) -> str:
    """Find a requested column with exact, normalized, or salary alias matching."""
    candidates = [requested] if requested else []
    candidates.extend(SALARY_ALIASES)

    for candidate in candidates:
        if not candidate:
            continue
        if candidate in headers:
            return candidate

    normalized_headers = {normalize_key(header): header for header in headers}
    for candidate in candidates:
        if not candidate:
            continue
        normalized = normalize_key(candidate)
        if normalized in normalized_headers:
            return normalized_headers[normalized]

    for candidate in candidates:
        if not candidate:
            continue
        normalized = normalize_key(candidate)
        for header in headers:
            header_key = normalize_key(header)
            if normalized and (normalized in header_key or header_key in normalized):
                return header

    available = ", ".join(headers)
    if requested:
        raise ValueError(f"Column not found: {requested}. Available: {available}")
    raise ValueError(f"No salary-like column found. Available: {available}")


def parse_number(value: Any) -> float | None:
    """Parse numbers from numeric cells or currency-like strings."""
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    text = text.replace("\uffe5", "").replace("\u00a5", "")
    text = text.replace("\u5143", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0))


def top_rows(
    records: list[dict[str, Any]],
    sort_column: str,
    limit: int,
    descending: bool,
) -> list[dict[str, Any]]:
    """Return top rows sorted by a numeric column."""
    sortable: list[tuple[float, dict[str, Any]]] = []
    for record in records:
        number = parse_number(record.get(sort_column))
        if number is None:
            continue
        sortable.append((number, record))

    sortable.sort(key=lambda item: item[0], reverse=descending)
    return [
        {**json_safe_record(record), "_sort_value": number}
        for number, record in sortable[:limit]
    ]


def json_safe_record(record: dict[str, Any]) -> dict[str, Any]:
    return {key: json_safe_value(value) for key, value in record.items()}


def json_safe_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Read and query an Excel workbook.")
    parser.add_argument("--file", required=True, help="Excel workbook path")
    parser.add_argument("--sheet", help="Worksheet name. Defaults to active sheet.")
    parser.add_argument("--top-by", help="Column to rank by, for example: \u6708\u85aa")
    parser.add_argument(
        "--limit",
        type=int,
        default=DEFAULT_LIMIT,
        help=f"Maximum number of rows to return. Defaults to {DEFAULT_LIMIT}.",
    )
    parser.add_argument(
        "--ascending",
        action="store_true",
        help="Sort ascending instead of descending.",
    )
    parser.add_argument(
        "--list-columns",
        action="store_true",
        help="Only list workbook sheets and columns.",
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    try:
        path = resolve_file_path(args.file)
        sheet_title, rows = load_sheet(path, args.sheet)
        headers, records = build_records(rows)

        if args.list_columns:
            result = {
                "file": str(path),
                "sheet": sheet_title,
                "columns": headers,
                "row_count": len(records),
            }
        else:
            if args.limit <= 0:
                raise ValueError("--limit must be a positive integer")
            sort_column = find_column(headers, args.top_by)
            result = {
                "file": str(path),
                "sheet": sheet_title,
                "sort_column": sort_column,
                "sort_order": "ascending" if args.ascending else "descending",
                "row_count": len(records),
                "rows": top_rows(
                    records,
                    sort_column=sort_column,
                    limit=args.limit,
                    descending=not args.ascending,
                ),
            }
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        raise SystemExit(1) from exc

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
