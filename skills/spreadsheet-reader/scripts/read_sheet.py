import argparse
import csv
import json
from pathlib import Path


def read_csv(path: Path, limit: int):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    return {"type": "csv", "columns": rows[0].keys() if rows else [], "rows": len(rows), "sample": rows[:limit]}


def read_excel(path: Path, limit: int):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "Missing dependency: openpyxl. Run pip install -r requirements.txt"}

    wb = load_workbook(path, read_only=True, data_only=True)
    result = {"type": "excel", "sheets": []}
    for ws in wb.worksheets:
        values = list(ws.iter_rows(values_only=True))
        headers = [str(v) if v is not None else "" for v in (values[0] if values else [])]
        sample = []
        for row in values[1 : limit + 1]:
            sample.append({headers[i] or f"column_{i+1}": row[i] for i in range(min(len(headers), len(row)))})
        result["sheets"].append({
            "name": ws.title,
            "rows": max(ws.max_row - 1, 0),
            "columns": headers,
            "sample": sample,
        })
    return result


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("path")
    parser.add_argument("--limit", type=int, default=8)
    args = parser.parse_args()

    path = Path(args.path.lstrip("/"))
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}, ensure_ascii=False, indent=2))
        return

    if path.suffix.lower() == ".csv":
        data = read_csv(path, args.limit)
    else:
        data = read_excel(path, args.limit)
    print(json.dumps(data, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
